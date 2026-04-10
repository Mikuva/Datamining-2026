#!/usr/bin/env python3
from __future__ import annotations

"""Notebook-friendly profiling of CZSO shortlist datasets.

This script builds on top of `pyczso.py` and `pyczso_lkod_bridge.py`.
It resolves the 165-row shortlist, downloads each dataset (prefer LKOD/VDB,
fall back to DataStat when needed), and for each dataset computes a small set of
practical characteristics:

- head() preview
- row count / column count
- detected time columns
- detected years contained in the dataset
- frequency and date range from the shortlist/catalogue metadata

The module can be used in two ways:
1) from the command line to save CSV / Markdown outputs,
2) from Jupyter to display each dataset in a notebook-friendly format.
"""

import argparse
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
import re
import sys
import time
from typing import Any, Iterable, Sequence
import unicodedata
import zipfile

import pandas as pd

THIS_DIR = Path(__file__).resolve().parent
if str(THIS_DIR) not in sys.path:
    sys.path.insert(0, str(THIS_DIR))

import pyczso  # noqa: E402
from pyczso_lkod_bridge import LKODBridge, detect_encoding  # noqa: E402

try:
    from IPython.display import Markdown, display  # type: ignore
except Exception:  # pragma: no cover
    Markdown = None
    display = None

PRIORITY_RANK = {"A": 0, "B": 1, "C": 2, "D": 3, "E": 4}
YEAR_RE = re.compile(r"(?<!\d)(18\d{2}|19\d{2}|20\d{2}|2100)(?!\d)")


@dataclass
class DatasetProfile:
    order: int
    priority: str
    rodina_datasetu: str
    dataset_id: str | None
    dataset_title: str | None
    source_kind: str | None
    source_code: str | None
    source_title: str | None
    selected_file: str | None
    selected_sheet: str | None
    selected_format: str | None
    status: str
    row_count: int | None = None
    column_count: int | None = None
    column_names: str | None = None
    time_columns: str | None = None
    years_min: int | None = None
    years_max: int | None = None
    years_count: int | None = None
    years_sample: str | None = None
    years_source: str | None = None
    frequency_catalog: str | None = None
    start_catalog: str | None = None
    end_catalog: str | None = None
    hlavni_skupina: str | None = None
    detailni_skupina: str | None = None
    uzemni_uroven: str | None = None
    zdroj_typ: str | None = None
    page: str | None = None
    preview_csv: str | None = None
    error: str | None = None


@dataclass
class TableAnalysis:
    head_df: pd.DataFrame | None
    row_count: int | None
    column_count: int | None
    column_names: list[str]
    time_columns: list[str]
    years: list[int]
    years_source: str | None
    selected_file: str | None
    selected_sheet: str | None
    selected_format: str | None
    notes: str | None = None


# ------------------------------
# Generic helpers
# ------------------------------


def normalize_name(value: Any) -> str:
    text = str(value)
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    return re.sub(r"[^a-z0-9]+", "", text)



def clean_string(value: Any) -> str | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None



def sanitize_filename(value: str, max_len: int = 90) -> str:
    text = unicodedata.normalize("NFKD", value)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"\s+", "_", text.strip())
    text = re.sub(r"[^0-9A-Za-z_\-]+", "", text)
    text = text.strip("._-")
    if not text:
        text = "dataset"
    return text[:max_len]



def guess_column(df: pd.DataFrame, candidates: Sequence[str]) -> str | None:
    if df is None or df.empty:
        return None
    lowered = {str(col).lower(): col for col in df.columns}
    for candidate in candidates:
        key = str(candidate).lower()
        if key in lowered:
            return lowered[key]
    normalized = {normalize_name(col): col for col in df.columns}
    for candidate in candidates:
        key = normalize_name(candidate)
        if key in normalized:
            return normalized[key]
    for col in df.columns:
        compact = normalize_name(col)
        for candidate in candidates:
            key = normalize_name(candidate)
            if key and key in compact:
                return col
    return None



def format_dt(value: Any) -> str | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    dt = pd.to_datetime(value, errors="coerce")
    if pd.isna(dt):
        text = clean_string(value)
        return text
    return dt.strftime("%Y-%m-%d")



def years_from_strings(values: Iterable[Any]) -> list[int]:
    years: set[int] = set()
    for value in values:
        if value is None:
            continue
        try:
            if pd.isna(value):
                continue
        except Exception:
            pass
        if isinstance(value, (int,)):
            if 1800 <= int(value) <= 2100:
                years.add(int(value))
                continue
        if isinstance(value, float):
            if value.is_integer() and 1800 <= int(value) <= 2100:
                years.add(int(value))
                continue
        if hasattr(value, "year") and not isinstance(value, str):
            try:
                year = int(value.year)
                if 1800 <= year <= 2100:
                    years.add(year)
                    continue
            except Exception:
                pass
        text = str(value)
        for match in YEAR_RE.findall(text):
            year = int(match)
            if 1800 <= year <= 2100:
                years.add(year)
    return sorted(years)



def detect_time_columns_from_head(df: pd.DataFrame) -> list[str]:
    if df is None or df.empty:
        return []

    name_patterns = [
        "rok", "year", "obdobi", "obdobi", "obdobi", "datum", "date",
        "cas", "time", "period", "mesic", "měsíc", "ctvrtleti", "čtvrtletí",
        "referencni", "reference", "obdobi", "week", "tyden", "týden",
    ]

    time_cols: list[str] = []
    for col in df.columns:
        ncol = normalize_name(col)
        if any(normalize_name(pat) in ncol for pat in name_patterns):
            time_cols.append(str(col))
            continue

        series = df[col].dropna().head(20)
        if series.empty:
            continue
        parsed_years = years_from_strings(series.tolist())
        if len(parsed_years) >= 2:
            time_cols.append(str(col))
            continue
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            time_cols.append(str(col))
            continue

    return list(dict.fromkeys(time_cols))



def format_years_sample(years: Sequence[int], max_items: int = 12) -> str | None:
    years = sorted(dict.fromkeys(int(y) for y in years))
    if not years:
        return None
    if len(years) <= max_items:
        return ", ".join(str(y) for y in years)
    front = max_items // 2
    back = max_items - front
    return ", ".join(str(y) for y in years[:front]) + " ... " + ", ".join(str(y) for y in years[-back:])



def dataframe_to_pretty_string(df: pd.DataFrame) -> str:
    with pd.option_context(
        "display.max_rows", None,
        "display.max_columns", None,
        "display.width", 240,
        "display.max_colwidth", 80,
    ):
        return df.to_string(index=False)


# ------------------------------
# Reading shortlist / catalogue
# ------------------------------


def read_table(path: Path, *, sheet_name: str | None = None) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(path)
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        if sheet_name is None:
            raise ValueError("Pro Excel je potřeba zadat sheet_name.")
        return pd.read_excel(path, sheet_name=sheet_name)
    raise ValueError(f"Nepodporovaný typ vstupu: {path}")



def sort_key_for_latest(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in ["end", "modified", "start"]:
        if col in out.columns:
            out[col] = pd.to_datetime(out[col], errors="coerce")
        else:
            out[col] = pd.NaT
    out["_end_sort"] = out["end"].fillna(pd.Timestamp("1900-01-01"))
    out["_modified_sort"] = out["modified"].fillna(pd.Timestamp("1900-01-01"))
    out["_start_sort"] = out["start"].fillna(pd.Timestamp("1900-01-01"))
    out["_dataset_sort"] = out["dataset_id"].astype(str)
    return out



def resolve_latest_dataset_per_family(catalogue: pd.DataFrame) -> pd.DataFrame:
    required = {"rodina_datasetu", "dataset_id"}
    missing = required - set(catalogue.columns)
    if missing:
        raise ValueError(f"Katalog musí obsahovat {sorted(required)}. Chybí: {sorted(missing)}")

    cat = sort_key_for_latest(catalogue)
    latest = (
        cat.sort_values(
            ["rodina_datasetu", "_end_sort", "_modified_sort", "_start_sort", "_dataset_sort"],
            ascending=[True, False, False, False, False],
        )
        .drop_duplicates(subset=["rodina_datasetu"], keep="first")
        .copy()
    )
    return latest.drop(columns=["_end_sort", "_modified_sort", "_start_sort", "_dataset_sort"])



def resolve_ranked_list(source_path: Path, *, shortlist_sheet: str, catalog_sheet: str) -> pd.DataFrame:
    shortlist = read_table(source_path, sheet_name=shortlist_sheet).copy()
    shortlist["_sheet_order"] = range(len(shortlist))

    if "priorita_shortlistu" not in shortlist.columns:
        shortlist["priorita_shortlistu"] = "Z"
    shortlist["priorita_shortlistu"] = shortlist["priorita_shortlistu"].fillna("Z").astype(str)
    shortlist["_priority_rank"] = shortlist["priorita_shortlistu"].map(PRIORITY_RANK).fillna(999)

    if "dataset_id" not in shortlist.columns:
        catalogue = read_table(source_path, sheet_name=catalog_sheet).copy()
        latest = resolve_latest_dataset_per_family(catalogue)
        join_cols = [
            col for col in [
                "rodina_datasetu", "dataset_id", "title", "dataset_iri", "page", "modified", "start", "end",
                "hlavni_skupina", "detailni_skupina", "relevance_pro_IKZ", "uzemni_uroven", "zdroj_typ",
                "periodicita_text", "periodicity",
            ] if col in latest.columns
        ]
        shortlist = shortlist.merge(latest[join_cols], on="rodina_datasetu", how="left", suffixes=("", "_catalog"))

    if "dataset_title" not in shortlist.columns:
        shortlist["dataset_title"] = shortlist["title"] if "title" in shortlist.columns else shortlist.get("rodina_datasetu")

    shortlist["dataset_id"] = shortlist["dataset_id"].map(clean_string)
    shortlist["dataset_title"] = shortlist["dataset_title"].map(clean_string)
    shortlist["rodina_datasetu"] = shortlist["rodina_datasetu"].map(clean_string)
    shortlist = shortlist.sort_values(["_priority_rank", "_sheet_order"], ascending=[True, True]).reset_index(drop=True)
    shortlist["poradi"] = range(1, len(shortlist) + 1)
    return shortlist


# ------------------------------
# DataStat helpers
# ------------------------------


def choose_selection(selections: pd.DataFrame) -> tuple[str | None, str | None, int]:
    if selections is None or selections.empty:
        return None, None, 0
    df = selections.copy()
    code_col = guess_column(df, ["kodVyberu", "kod_vyberu", "kod", "selection_code", "vyberKod", "kodvyberu"])
    title_col = guess_column(df, ["nazevVyberu", "nazev_vyberu", "nazev", "title", "selection_title", "nazevvyberu"])
    default_col = guess_column(df, ["vychozi", "jeVychozi", "default", "defaultni"])
    version_col = guess_column(df, ["verzeVyberu", "verze", "version"])
    modified_col = guess_column(df, ["datumAktualizace", "modified", "updated", "zmeneno", "casZmenyDefinice"])
    if code_col is None:
        return None, None, len(df)

    if default_col is not None:
        df["_default_rank"] = (
            df[default_col].fillna(False).astype(str).str.strip().str.lower().isin(["true", "1", "ano", "yes"]).astype(int)
        )
    else:
        df["_default_rank"] = 0
    if version_col is not None:
        df["_version_rank"] = pd.to_numeric(df[version_col], errors="coerce").fillna(-1)
    else:
        df["_version_rank"] = -1
    if modified_col is not None:
        df["_modified_rank"] = pd.to_datetime(df[modified_col], errors="coerce").fillna(pd.Timestamp("1900-01-01"))
    else:
        df["_modified_rank"] = pd.Timestamp("1900-01-01")
    if title_col is not None:
        text = df[title_col].fillna("").astype(str).str.lower()
        df["_title_rank"] = (
            text.str.contains(r"\b(celkem|vse|vše|total)\b", regex=True).astype(int) * 3
            + text.str.contains(r"\b(obec|obce|kraj|kraje)\b", regex=True).astype(int)
            - text.str.contains(r"\b(test|archiv|historie)\b", regex=True).astype(int)
        )
    else:
        df["_title_rank"] = 0
    best = df.sort_values(["_default_rank", "_title_rank", "_version_rank", "_modified_rank"], ascending=[False, False, False, False]).iloc[0]
    code = clean_string(best[code_col])
    title = clean_string(best[title_col]) if title_col is not None else None
    return code, title, len(df)


# ------------------------------
# CSV / XLSX / ZIP profiling
# ------------------------------


def read_bytes_head(path: Path, size: int = 200000) -> bytes:
    with path.open('rb') as f:
        return f.read(size)


def coalesce_non_na(*values: Any) -> Any:
    for value in values:
        try:
            if pd.isna(value):
                continue
        except Exception:
            pass
        if value is not None:
            return value
    return None


def _csv_read_attempts(path: Path, **kwargs: Any):
    raw = read_bytes_head(path)
    last_exc: Exception | None = None
    for enc in detect_encoding(raw):
        try:
            return pd.read_csv(path, encoding=enc, sep=None, engine="python", **kwargs)
        except Exception as exc:
            last_exc = exc
    raise RuntimeError(f"CSV reading failed for {path}: {last_exc}")



def read_csv_head(path: Path, nrows: int) -> pd.DataFrame:
    return _csv_read_attempts(path, nrows=nrows)



def count_csv_rows(path: Path, chunksize: int = 50000) -> int:
    total = 0
    raw = read_bytes_head(path)
    last_exc: Exception | None = None
    for enc in detect_encoding(raw):
        try:
            for chunk in pd.read_csv(path, encoding=enc, sep=None, engine="python", chunksize=chunksize):
                total += len(chunk)
            return int(total)
        except Exception as exc:
            total = 0
            last_exc = exc
    raise RuntimeError(f"CSV row count failed for {path}: {last_exc}")



def detect_years_from_csv(path: Path, head_df: pd.DataFrame, chunksize: int = 50000) -> tuple[list[int], list[str], str | None]:
    header_years = years_from_strings(list(head_df.columns))
    time_cols = detect_time_columns_from_head(head_df)
    if not time_cols:
        if header_years:
            return header_years, [], "header_columns"
        return [], [], None

    years: set[int] = set()
    raw = read_bytes_head(path)
    last_exc: Exception | None = None
    for enc in detect_encoding(raw):
        try:
            for chunk in pd.read_csv(path, encoding=enc, sep=None, engine="python", usecols=time_cols, chunksize=chunksize):
                for col in time_cols:
                    if col in chunk.columns:
                        years.update(years_from_strings(chunk[col].dropna().tolist()))
            break
        except Exception as exc:
            years.clear()
            last_exc = exc
            continue
    if years:
        return sorted(years), time_cols, "data_columns"
    if header_years:
        return header_years, time_cols, "header_columns"
    if last_exc is not None:
        return [], time_cols, f"time_columns_detected_but_scan_failed: {last_exc}"
    return [], time_cols, None



def analyze_csv_file(path: Path, head_rows: int) -> TableAnalysis:
    head_df = read_csv_head(path, nrows=head_rows)
    row_count = count_csv_rows(path)
    years, time_cols, years_source = detect_years_from_csv(path, head_df)
    return TableAnalysis(
        head_df=head_df,
        row_count=row_count,
        column_count=int(len(head_df.columns)),
        column_names=[str(c) for c in head_df.columns],
        time_columns=time_cols,
        years=years,
        years_source=years_source,
        selected_file=str(path.name),
        selected_sheet=None,
        selected_format="csv",
        notes=None,
    )



def analyze_excel_file(path: Path, head_rows: int) -> TableAnalysis:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("Pro práci s Excel soubory je potřeba openpyxl.") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        selected_sheet = None
        row_count = None
        column_count = None
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row = int(ws.max_row or 0)
            max_col = int(ws.max_column or 0)
            if max_row > 0 and max_col > 0:
                selected_sheet = sheet_name
                row_count = max(0, max_row - 1)
                column_count = max_col
                break
        if selected_sheet is None:
            raise RuntimeError(f"Excel soubor {path.name} neobsahuje čitelný list.")
    finally:
        wb.close()

    head_df = pd.read_excel(path, sheet_name=selected_sheet, nrows=head_rows)
    time_cols = detect_time_columns_from_head(head_df)
    years = years_from_strings(list(head_df.columns))
    years_source = "header_columns" if years else None

    if time_cols:
        try:
            full_time = pd.read_excel(path, sheet_name=selected_sheet, usecols=time_cols)
            extracted: set[int] = set()
            for col in time_cols:
                if col in full_time.columns:
                    extracted.update(years_from_strings(full_time[col].dropna().tolist()))
            if extracted:
                years = sorted(extracted)
                years_source = "data_columns"
        except Exception:
            pass

    return TableAnalysis(
        head_df=head_df,
        row_count=row_count,
        column_count=int(column_count or len(head_df.columns)),
        column_names=[str(c) for c in head_df.columns],
        time_columns=time_cols,
        years=sorted(dict.fromkeys(years)),
        years_source=years_source,
        selected_file=str(path.name),
        selected_sheet=selected_sheet,
        selected_format=path.suffix.lower().lstrip("."),
        notes=None,
    )



def choose_archive_table_file(files: list[Path]) -> Path:
    def score(path: Path) -> tuple[int, int, int, str]:
        name = normalize_name(path.name)
        ext_score = 100 if path.suffix.lower() == ".csv" else 60 if path.suffix.lower() in {".xlsx", ".xls", ".xlsm"} else 0
        penalty = 0
        for bad in ["readme", "schema", "meta", "ciselnik", "codebook", "documentation", "popis"]:
            if bad in name:
                penalty -= 30
        size_score = int(min(path.stat().st_size / 1_000_000, 200)) if path.exists() else 0
        return (ext_score + penalty + size_score, size_score, -len(path.parts), path.name)

    return sorted(files, key=score, reverse=True)[0]



def analyze_zip_file(path: Path, head_rows: int) -> TableAnalysis:
    extract_dir = path.parent / f"extract_{path.stem}"
    extract_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "r") as zf:
        zf.extractall(extract_dir)
    candidates = [p for p in extract_dir.rglob("*") if p.is_file() and p.suffix.lower() in {".csv", ".xlsx", ".xls", ".xlsm"}]
    if not candidates:
        return TableAnalysis(
            head_df=None,
            row_count=None,
            column_count=None,
            column_names=[],
            time_columns=[],
            years=[],
            years_source=None,
            selected_file=None,
            selected_sheet=None,
            selected_format="zip",
            notes="Archiv neobsahuje CSV/XLSX soubor, který by šel automaticky profilovat.",
        )
    chosen = choose_archive_table_file(candidates)
    if chosen.suffix.lower() == ".csv":
        analysis = analyze_csv_file(chosen, head_rows=head_rows)
    else:
        analysis = analyze_excel_file(chosen, head_rows=head_rows)
    note = f"Archiv obsahuje {len(candidates)} čitelných souborů; profilován byl vybraný soubor `{chosen.name}`."
    if analysis.notes:
        note = analysis.notes + " | " + note
    analysis.notes = note
    analysis.selected_format = f"zip->{analysis.selected_format}"
    return analysis



def analyze_dataframe(df: pd.DataFrame, head_rows: int, *, selected_format: str, selected_file: str | None = None, selected_sheet: str | None = None) -> TableAnalysis:
    head_df = df.head(head_rows).copy()
    time_cols = detect_time_columns_from_head(head_df)
    years = years_from_strings(list(df.columns))
    years_source = "header_columns" if years else None
    if time_cols:
        extracted: set[int] = set()
        for col in time_cols:
            if col in df.columns:
                extracted.update(years_from_strings(df[col].dropna().tolist()))
        if extracted:
            years = sorted(extracted)
            years_source = "data_columns"
    return TableAnalysis(
        head_df=head_df,
        row_count=int(len(df)),
        column_count=int(len(df.columns)),
        column_names=[str(c) for c in df.columns],
        time_columns=time_cols,
        years=years,
        years_source=years_source,
        selected_file=selected_file,
        selected_sheet=selected_sheet,
        selected_format=selected_format,
        notes=None,
    )


# ------------------------------
# Fetch and profile one dataset
# ------------------------------


def fetch_and_profile_dataset(
    datastat_client: pyczso.CZSOClient,
    lkod_bridge: LKODBridge,
    dataset_id: str,
    *,
    head_rows: int,
    download_dir: Path,
    force_redownload: bool,
) -> tuple[TableAnalysis | None, str | None, str | None, str | None, str, str | None]:
    lkod_error: str | None = None

    try:
        pointer = lkod_bridge.download_resource(dataset_id, dest_dir=download_dir, force_redownload=force_redownload)
        local_path = Path(pointer["download_path"])
        ext = str(pointer.get("extension") or local_path.suffix.lstrip(".")).lower()
        source_title = clean_string(pointer.get("title")) or clean_string(pointer.get("format"))
        if ext == "csv":
            analysis = analyze_csv_file(local_path, head_rows=head_rows)
        elif ext in {"xlsx", "xls", "xlsm"}:
            analysis = analyze_excel_file(local_path, head_rows=head_rows)
        elif ext == "zip":
            analysis = analyze_zip_file(local_path, head_rows=head_rows)
        else:
            return None, "LKOD", f"resource#{pointer.get('resource_num')}", source_title, "UNSUPPORTED_FORMAT", f"Stažený formát `{ext}` není zatím podporovaný pro automatické profilování."
        return analysis, "LKOD", f"resource#{pointer.get('resource_num')}", source_title, "OK_LKOD", analysis.notes
    except Exception as exc:
        lkod_error = str(exc)

    try:
        selections = datastat_client.get_dataset_selections(dataset_id, as_frame=True)
    except Exception as exc:
        msg = f"LKOD/VDB selhalo: {lkod_error} | DataStat metadata selhala: {exc}" if lkod_error else f"DataStat metadata selhala: {exc}"
        return None, None, None, None, "ERROR_SELECTIONS", msg

    selection_code, selection_title, selection_count = choose_selection(selections)
    if not selection_code:
        msg = f"Nebyl nalezen použitelný výběr. Počet nalezených výběrů: {selection_count}"
        if lkod_error:
            msg = f"LKOD/VDB selhalo: {lkod_error} | {msg}"
        return None, "DATASTAT", None, None, "NO_SELECTION", msg

    try:
        df = datastat_client.get_table(selection_code, format="CSV")
    except Exception as exc:
        msg = f"Stažení DataStat výběru selhalo: {exc}"
        if lkod_error:
            msg = f"LKOD/VDB selhalo: {lkod_error} | {msg}"
        return None, "DATASTAT", selection_code, selection_title, "ERROR_TABLE", msg
    if not isinstance(df, pd.DataFrame):
        msg = "API nevrátilo DataFrame pro CSV formát."
        if lkod_error:
            msg = f"LKOD/VDB selhalo: {lkod_error} | {msg}"
        return None, "DATASTAT", selection_code, selection_title, "ERROR_TABLE", msg

    analysis = analyze_dataframe(df, head_rows=head_rows, selected_format="datastat_csv", selected_file=f"selection_{selection_code}.csv")
    return analysis, "DATASTAT", selection_code, selection_title, "OK_DATASTAT", lkod_error


# ------------------------------
# Reporting / display
# ------------------------------


def ensure_output_dirs(base_dir: Path) -> tuple[Path, Path, Path]:
    heads_dir = base_dir / "heads_csv"
    base_dir.mkdir(parents=True, exist_ok=True)
    heads_dir.mkdir(parents=True, exist_ok=True)
    downloads_dir = base_dir / "_downloads"
    downloads_dir.mkdir(parents=True, exist_ok=True)
    return base_dir, heads_dir, downloads_dir



def build_profile_record(
    row: pd.Series,
    analysis: TableAnalysis | None,
    *,
    order: int,
    priority: str,
    family: str,
    dataset_id: str | None,
    dataset_title: str | None,
    source_kind: str | None,
    source_code: str | None,
    source_title: str | None,
    status: str,
    preview_csv: str | None,
    error: str | None,
) -> DatasetProfile:
    years = analysis.years if analysis is not None else []
    column_names = ", ".join((analysis.column_names[:15] if analysis else []))
    if analysis and len(analysis.column_names) > 15:
        column_names += " ..."

    freq = clean_string(row.get("periodicita_text"))
    periodicity_raw = clean_string(row.get("periodicity"))
    if periodicity_raw:
        freq = f"{freq} [{periodicity_raw}]" if freq else periodicity_raw

    profile = DatasetProfile(
        order=order,
        priority=priority,
        rodina_datasetu=family,
        dataset_id=dataset_id,
        dataset_title=dataset_title,
        source_kind=source_kind,
        source_code=source_code,
        source_title=source_title,
        selected_file=analysis.selected_file if analysis else None,
        selected_sheet=analysis.selected_sheet if analysis else None,
        selected_format=analysis.selected_format if analysis else None,
        status=status,
        row_count=analysis.row_count if analysis else None,
        column_count=analysis.column_count if analysis else None,
        column_names=column_names or None,
        time_columns=", ".join(analysis.time_columns) if analysis and analysis.time_columns else None,
        years_min=min(years) if years else None,
        years_max=max(years) if years else None,
        years_count=len(years) if years else None,
        years_sample=format_years_sample(years),
        years_source=analysis.years_source if analysis else None,
        frequency_catalog=freq,
        start_catalog=format_dt(coalesce_non_na(row.get("start"), row.get("nejstarsi_start"))),
        end_catalog=format_dt(coalesce_non_na(row.get("end"), row.get("nejnovejsi_end"))),
        hlavni_skupina=clean_string(row.get("hlavni_skupina")),
        detailni_skupina=clean_string(row.get("detailni_skupina")),
        uzemni_uroven=clean_string(row.get("uzemni_uroven")),
        zdroj_typ=clean_string(row.get("zdroj_typ")),
        page=clean_string(row.get("page")),
        preview_csv=preview_csv,
        error=error,
    )
    return profile



def profile_to_series(profile: DatasetProfile) -> pd.DataFrame:
    preferred_order = [
        "dataset_id", "dataset_title", "status", "source_kind", "source_code", "source_title",
        "selected_file", "selected_sheet", "selected_format", "row_count", "column_count",
        "time_columns", "years_min", "years_max", "years_count", "years_sample", "years_source",
        "frequency_catalog", "start_catalog", "end_catalog", "hlavni_skupina", "detailni_skupina",
        "uzemni_uroven", "zdroj_typ", "page", "preview_csv", "error",
    ]
    data = asdict(profile)
    ordered = {k: data.get(k) for k in preferred_order}
    return pd.DataFrame({"hodnota": ordered})



def append_markdown_entry(report_path: Path, row: pd.Series, profile: DatasetProfile, preview_df: pd.DataFrame | None) -> None:
    lines = [f"## {profile.order:03d}. [{profile.priority}] {profile.rodina_datasetu}", ""]
    for key, value in profile_to_series(profile)["hodnota"].items():
        if value is not None:
            lines.append(f"- {key}: {value}")
    lines.append("")
    if preview_df is not None and not preview_df.empty:
        lines.extend(["```text", dataframe_to_pretty_string(preview_df), "```", ""])
    with report_path.open("a", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")



def display_profile_in_notebook(profile: DatasetProfile, preview_df: pd.DataFrame | None) -> None:
    if display is None or Markdown is None:
        print(f"{profile.order:03d}. [{profile.priority}] {profile.rodina_datasetu}")
        print(profile_to_series(profile))
        if preview_df is not None:
            print(preview_df)
        return

    display(Markdown(f"## {profile.order:03d}. [{profile.priority}] {profile.rodina_datasetu}"))
    display(profile_to_series(profile))
    if preview_df is not None and not preview_df.empty:
        display(preview_df)



def print_profile_to_console(profile: DatasetProfile, preview_df: pd.DataFrame | None) -> None:
    print("=" * 100)
    print(f"{profile.order:03d}. [{profile.priority}] {profile.rodina_datasetu}")
    print(f"dataset_id      : {profile.dataset_id}")
    print(f"status          : {profile.status}")
    if profile.source_kind:
        print(f"source_kind     : {profile.source_kind}")
    if profile.source_code:
        print(f"source_code     : {profile.source_code}")
    if profile.selected_file:
        print(f"selected_file   : {profile.selected_file}")
    if profile.row_count is not None:
        print(f"row_count       : {profile.row_count}")
    if profile.column_count is not None:
        print(f"column_count    : {profile.column_count}")
    if profile.years_sample:
        print(f"years_sample    : {profile.years_sample}")
    if profile.error:
        print(f"error           : {profile.error}")
    if preview_df is not None and not preview_df.empty:
        print("-" * 100)
        print(dataframe_to_pretty_string(preview_df))
    print()


# ------------------------------
# Public orchestration API
# ------------------------------


def run_shortlist_profiles(
    *,
    catalog_xlsx: str | Path,
    output_dir: str | Path = "czso_dataset_profiles",
    shortlist_sheet: str = "Shortlist_IKZ",
    catalog_sheet: str = "Katalog_trideny",
    head_rows: int = 5,
    max_datasets: int | None = None,
    priority: Sequence[str] | None = None,
    pause_seconds: float = 0.0,
    force_redownload: bool = False,
    dry_run: bool = False,
    display_mode: str = "none",
) -> pd.DataFrame:
    source_path = Path(catalog_xlsx).expanduser().resolve()
    out_dir = Path(output_dir).expanduser().resolve()
    out_dir, heads_dir, downloads_dir = ensure_output_dirs(out_dir)

    ranked = resolve_ranked_list(source_path, shortlist_sheet=shortlist_sheet, catalog_sheet=catalog_sheet)
    if priority:
        allowed = {str(p).strip().upper() for p in priority}
        ranked = ranked[ranked["priorita_shortlistu"].str.upper().isin(allowed)].reset_index(drop=True)
        ranked["poradi"] = range(1, len(ranked) + 1)
    if max_datasets is not None:
        ranked = ranked.head(max_datasets).copy()
        ranked["poradi"] = range(1, len(ranked) + 1)

    resolved_list_path = out_dir / "resolved_ranked_list.csv"
    ranked.to_csv(resolved_list_path, index=False, encoding="utf-8-sig")

    report_path = out_dir / "dataset_profiles_report.md"
    report_header = (
        "# Profil CZSO shortlistu\n\n"
        f"- generováno: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"- zdroj seznamu: `{source_path.name}` / sheet `{shortlist_sheet}`\n"
        f"- počet datasetů: {len(ranked)}\n"
        f"- head rows: {head_rows}\n\n"
    )
    report_path.write_text(report_header, encoding="utf-8")

    manifest: list[dict[str, Any]] = []

    if dry_run:
        for _, row in ranked.iterrows():
            family = clean_string(row.get("rodina_datasetu")) or "N/A"
            profile = build_profile_record(
                row, None,
                order=int(row["poradi"]),
                priority=str(row["priorita_shortlistu"]),
                family=family,
                dataset_id=clean_string(row.get("dataset_id")),
                dataset_title=clean_string(row.get("dataset_title")) or family,
                source_kind=None,
                source_code=None,
                source_title=None,
                status="DRY_RUN",
                preview_csv=None,
                error=None,
            )
            manifest.append(asdict(profile))
            append_markdown_entry(report_path, row, profile, preview_df=None)
        profiles_df = pd.DataFrame(manifest)
        profiles_df.to_csv(out_dir / "dataset_profiles_manifest.csv", index=False, encoding="utf-8-sig")
        return profiles_df

    datastat_client = pyczso.CZSOClient()
    lkod_bridge = LKODBridge()
    try:
        for _, row in ranked.iterrows():
            order = int(row["poradi"])
            priority_text = str(row["priorita_shortlistu"])
            family = clean_string(row.get("rodina_datasetu")) or "N/A"
            dataset_id = clean_string(row.get("dataset_id"))
            dataset_title = clean_string(row.get("dataset_title")) or family

            if not dataset_id:
                preview_df = None
                analysis = None
                source_kind = None
                source_code = None
                source_title = None
                status = "MISSING_DATASET_ID"
                error = "V shortlistu ani katalogu se nepodařilo dohledat dataset_id."
            else:
                analysis, source_kind, source_code, source_title, status, error = fetch_and_profile_dataset(
                    datastat_client,
                    lkod_bridge,
                    dataset_id,
                    head_rows=head_rows,
                    download_dir=downloads_dir,
                    force_redownload=force_redownload,
                )
                preview_df = None if analysis is None else analysis.head_df

            preview_csv_rel = None
            if preview_df is not None and not preview_df.empty:
                fname = f"{order:03d}_{sanitize_filename(dataset_id or 'na')}_{sanitize_filename(dataset_title)}.csv"
                preview_path = heads_dir / fname
                preview_df.to_csv(preview_path, index=False, encoding="utf-8-sig")
                preview_csv_rel = str(preview_path.relative_to(out_dir))

            profile = build_profile_record(
                row,
                analysis,
                order=order,
                priority=priority_text,
                family=family,
                dataset_id=dataset_id,
                dataset_title=dataset_title,
                source_kind=source_kind,
                source_code=source_code,
                source_title=source_title,
                status=status,
                preview_csv=preview_csv_rel,
                error=error,
            )
            manifest.append(asdict(profile))
            append_markdown_entry(report_path, row, profile, preview_df)

            if display_mode == "notebook":
                display_profile_in_notebook(profile, preview_df)
            elif display_mode == "console":
                print_profile_to_console(profile, preview_df)

            if pause_seconds:
                time.sleep(pause_seconds)
    finally:
        datastat_client.close()
        lkod_bridge.close()

    profiles_df = pd.DataFrame(manifest)
    profiles_df.to_csv(out_dir / "dataset_profiles_manifest.csv", index=False, encoding="utf-8-sig")
    return profiles_df



def display_shortlist_profiles(
    *,
    catalog_xlsx: str | Path,
    output_dir: str | Path = "czso_dataset_profiles",
    shortlist_sheet: str = "Shortlist_IKZ",
    catalog_sheet: str = "Katalog_trideny",
    head_rows: int = 5,
    max_datasets: int | None = None,
    priority: Sequence[str] | None = None,
    pause_seconds: float = 0.0,
    force_redownload: bool = False,
) -> pd.DataFrame:
    """Notebook-first wrapper.

    Use this directly from Jupyter. It displays each dataset one by one and also
    returns the manifest DataFrame.
    """
    return run_shortlist_profiles(
        catalog_xlsx=catalog_xlsx,
        output_dir=output_dir,
        shortlist_sheet=shortlist_sheet,
        catalog_sheet=catalog_sheet,
        head_rows=head_rows,
        max_datasets=max_datasets,
        priority=priority,
        pause_seconds=pause_seconds,
        force_redownload=force_redownload,
        dry_run=False,
        display_mode="notebook",
    )


# ------------------------------
# CLI
# ------------------------------


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Vytvoří head() + základní profil pro CZSO shortlist datasety.")
    parser.add_argument("--catalog-xlsx", required=True, help="Cesta k czso_katalog_trideny.xlsx")
    parser.add_argument("--shortlist-sheet", default="Shortlist_IKZ", help="Sheet se shortlistem. Default: Shortlist_IKZ")
    parser.add_argument("--catalog-sheet", default="Katalog_trideny", help="Sheet s plným katalogem. Default: Katalog_trideny")
    parser.add_argument("--output-dir", default="czso_dataset_profiles", help="Výstupní adresář")
    parser.add_argument("--head-rows", type=int, default=5, help="Kolik řádků vracet přes head().")
    parser.add_argument("--max-datasets", type=int, default=None, help="Volitelný limit počtu datasetů")
    parser.add_argument("--priority", nargs="*", default=None, help="Filtr priorit, např. --priority A B")
    parser.add_argument("--pause-seconds", type=float, default=0.0, help="Pauza mezi dotazy")
    parser.add_argument("--force-redownload", action="store_true", help="Vynutí opětovné stažení zdrojových souborů")
    parser.add_argument("--dry-run", action="store_true", help="Jen připraví pořadí a metadata bez stahování dat")
    parser.add_argument("--quiet", action="store_true", help="Nevypisovat průběh do konzole")
    return parser



def main() -> int:
    args = build_parser().parse_args()
    display_mode = "none" if args.quiet else "console"
    profiles_df = run_shortlist_profiles(
        catalog_xlsx=args.catalog_xlsx,
        output_dir=args.output_dir,
        shortlist_sheet=args.shortlist_sheet,
        catalog_sheet=args.catalog_sheet,
        head_rows=args.head_rows,
        max_datasets=args.max_datasets,
        priority=args.priority,
        pause_seconds=args.pause_seconds,
        force_redownload=args.force_redownload,
        dry_run=args.dry_run,
        display_mode=display_mode,
    )
    out_dir = Path(args.output_dir).expanduser().resolve()
    print("Hotovo.")
    print(f"- manifest      : {out_dir / 'dataset_profiles_manifest.csv'}")
    print(f"- report        : {out_dir / 'dataset_profiles_report.md'}")
    print(f"- resolved list : {out_dir / 'resolved_ranked_list.csv'}")
    print(f"- heads dir     : {out_dir / 'heads_csv'}")
    print(f"- dataset count : {len(profiles_df)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
